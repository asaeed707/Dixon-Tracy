from __future__ import annotations

import json
import os
import sys
import smtplib
from collections import defaultdict
from datetime import UTC, date, datetime, time, timedelta
from email.message import EmailMessage
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table as PdfTable, TableStyle

from geotab_connect import GeotabClient


LOCAL_TZ = ZoneInfo("America/Los_Angeles")
PROJECT_ROOT = Path(__file__).resolve().parents[1]
OUTPUT_DIR = Path(os.environ.get("OUTPUT_DIR", PROJECT_ROOT / "outputs"))
OUTPUT_AUDIT_DIR = Path(os.environ.get("OUTPUT_AUDIT_DIR", PROJECT_ROOT / "work"))
MIN_DURATION = timedelta(minutes=25)
MERGE_GAP = timedelta(minutes=25)
LONG_LOAD = timedelta(hours=1, minutes=30)
MAX_DURATION = timedelta(hours=4)
OPERATING_START = time(5, 0)
OPERATING_END = time(17, 0)
ZONES = [
    {"label": "BB Dixon", "title": "Dixon", "rule_token": "DIXON", "sheet": "BB Dixon"},
    {"label": "BB Tracy", "title": "Tracy", "rule_token": "TRACY", "sheet": "BB Tracy"},
]


def report_day() -> date:
    override = os.environ.get("REPORT_DATE")
    if override:
        return date.fromisoformat(override)
    today = datetime.now(LOCAL_TZ).date()
    if today.weekday() == 0:
        return today - timedelta(days=3)
    return today - timedelta(days=1)


REPORT_DAY = report_day()
OUTPUT_XLSX = OUTPUT_DIR / f"bb_dixon_tracy_loading_times_{REPORT_DAY.isoformat()}.xlsx"
OUTPUT_JSON = OUTPUT_AUDIT_DIR / f"bb_dixon_tracy_loading_times_{REPORT_DAY.isoformat()}.json"


def display_date(value: date) -> str:
    return f"{value.strftime('%B')} {value.day}"


OUTPUT_PDF = OUTPUT_DIR / f"Dixon-Tracy Loading Times for {display_date(REPORT_DAY)}, {REPORT_DAY.year}.pdf"


def require_env(name: str) -> str:
    value = os.environ.get(name)
    if not value:
        raise SystemExit(f"Missing required environment variable: {name}")
    return value


def iso_z(dt: datetime) -> str:
    return dt.astimezone(UTC).isoformat(timespec="milliseconds").replace("+00:00", "Z")


def parse_api_datetime(value: str | None) -> datetime | None:
    if not value:
        return None
    dt = datetime.fromisoformat(value.replace("Z", "+00:00"))
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=UTC)
    return dt.astimezone(LOCAL_TZ).replace(tzinfo=None)


def ref_id(value: object) -> str:
    if isinstance(value, dict):
        return str(value.get("id") or "")
    return str(value or "")


def fetch_lookup(client: GeotabClient, type_name: str, fields: list[str], limit: int = 50000) -> dict[str, dict]:
    rows = client.call(
        "Get",
        {
            "typeName": type_name,
            "credentials": client.credentials,
            "resultsLimit": limit,
            "propertySelector": {"fields": fields, "isIncluded": True},
        },
    )
    return {row["id"]: row for row in rows if row.get("id")}


def event_type(rule_name: str) -> str | None:
    upper_name = rule_name.upper()
    if "EXIT" in upper_name:
        return "exit"
    if "ENTRY" in upper_name or "ENTER" in upper_name:
        return "entry"
    return None


def fetch_inputs(client: GeotabClient) -> tuple[dict[str, list[dict]], dict[str, dict], dict[str, list[str]]]:
    start_local = datetime.combine(REPORT_DAY, time.min, LOCAL_TZ)
    end_local = start_local + timedelta(days=1)
    devices = fetch_lookup(client, "Device", ["id", "name"])
    rules = fetch_lookup(client, "Rule", ["id", "name"], limit=5000)
    events_by_zone: dict[str, list[dict]] = {}
    rule_names_by_zone: dict[str, list[str]] = {}

    for zone_config in ZONES:
        selected_rules = {
            rule_id: rule.get("name", "")
            for rule_id, rule in rules.items()
            if zone_config["rule_token"] in rule.get("name", "").upper()
            and event_type(rule.get("name", "")) in {"entry", "exit"}
        }
        selected_types = {event_type(name) for name in selected_rules.values()}
        if selected_types != {"entry", "exit"}:
            raise SystemExit(
                f"Could not find both ENTRY and EXIT rules for {zone_config['label']}. "
                f"Matched: {sorted(selected_rules.values())}"
            )

        raw_events = []
        for rule_id in selected_rules:
            raw_events.extend(
                client.call(
                    "Get",
                    {
                        "typeName": "ExceptionEvent",
                        "credentials": client.credentials,
                        "search": {
                            "fromDate": iso_z(start_local),
                            "toDate": iso_z(end_local),
                            "ruleSearch": {"id": rule_id},
                        },
                        "resultsLimit": 50000,
                    },
                )
            )

        events = []
        seen = set()
        for item in raw_events:
            timestamp = parse_api_datetime(item.get("activeFrom"))
            rule_id = ref_id(item.get("rule"))
            kind = event_type(selected_rules.get(rule_id, ""))
            device_id = ref_id(item.get("device"))
            dedupe_key = (device_id, timestamp, kind)
            if timestamp is None or kind is None or not device_id or dedupe_key in seen:
                continue
            seen.add(dedupe_key)
            events.append(
                {
                    "device": devices.get(device_id, {}).get("name") or device_id,
                    "device_id": device_id,
                    "timestamp": timestamp,
                    "type": kind,
                    "rule": selected_rules.get(rule_id, ""),
                    "event_id": item.get("id", ""),
                }
            )
        events.sort(key=lambda item: (item["timestamp"], item["device"], item["type"]))
        events_by_zone[zone_config["label"]] = events
        rule_names_by_zone[zone_config["label"]] = sorted(selected_rules.values())

    return events_by_zone, devices, rule_names_by_zone


def sessions_for_zone(events: list[dict]) -> tuple[list[dict], list[dict]]:
    ignored = []
    grouped = defaultdict(list)
    for event in events:
        grouped[event["device_id"]].append(event)

    sessions = []
    for group in grouped.values():
        group.sort(key=lambda item: item["timestamp"])
        current_entry = None
        index = 0
        while index < len(group):
            event = group[index]
            if current_entry is None:
                if event["type"] == "entry":
                    current_entry = event
                else:
                    ignored.append({**event, "reason": "Exit event with no open entry"})
                index += 1
                continue

            if event["type"] == "entry":
                ignored.append({**event, "reason": "Duplicate entry while prior entry is still open"})
                index += 1
                continue

            duration = event["timestamp"] - current_entry["timestamp"]
            next_event = group[index + 1] if index + 1 < len(group) else None
            immediate_reentry = (
                duration < MIN_DURATION
                and next_event is not None
                and next_event["type"] == "entry"
                and next_event["timestamp"] - event["timestamp"] < MERGE_GAP
            )
            if immediate_reentry:
                ignored.append({**event, "reason": "GPS blip exit under 25 minutes"})
                ignored.append({**next_event, "reason": "Immediate re-entry after GPS blip"})
                index += 2
                continue

            start = current_entry["timestamp"]
            end = event["timestamp"]
            session = {
                "device": current_entry["device"],
                "device_id": current_entry["device_id"],
                "start": start,
                "end": end,
                "duration": duration,
                "entry_event_id": current_entry["event_id"],
                "exit_event_id": event["event_id"],
            }
            current_entry = None
            index += 1

            if start.date() != REPORT_DAY or not OPERATING_START <= start.time() < OPERATING_END:
                ignored.append({**session, "reason": "Entry outside 5:00 AM-5:00 PM start window"})
                continue
            if duration > MAX_DURATION:
                ignored.append({**session, "reason": "Probable error: duration over 4 hours"})
            elif duration >= MIN_DURATION:
                sessions.append(session)
            else:
                ignored.append({**session, "reason": "Zone visit under 25 minutes"})

        if current_entry is not None:
            ignored.append({**current_entry, "reason": "Entry event with no matching exit"})

    sessions.sort(key=lambda item: item["start"])
    ignored.sort(key=lambda item: (item.get("timestamp") or item.get("start"), item["device"]))
    return sessions, ignored


def write_zone_sheet(wb: Workbook, sheet_name: str, sessions: list[dict]) -> None:
    ws = wb.create_sheet(sheet_name)
    ws.append(["Device", "Date", "Start Time", "Stop Time", "Duration"])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    yellow_fill = PatternFill("solid", fgColor="FFFF00")
    grid = Side(style="thin", color="D9D9D9")
    border = Border(left=grid, right=grid, top=grid, bottom=grid)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", size=14)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = border

    for session in sessions:
        ws.append(
            [
                session["device"],
                session["start"].date(),
                session["start"],
                session["end"],
                session["duration"],
            ]
        )

    first_data_row = 2
    last_data_row = 1 + len(sessions)
    average_row = last_data_row + 1
    ws.cell(average_row, 4, "Average")
    ws.cell(average_row, 5, timedelta(seconds=sum(item["duration"].total_seconds() for item in sessions) / len(sessions)) if sessions else None)

    if sessions:
        table_name = "".join(ch for ch in sheet_name if ch.isalnum()) + "LoadingTimes"
        table = Table(displayName=table_name, ref=f"A1:E{last_data_row}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=False, showColumnStripes=False)
        ws.add_table(table)

    for row in ws.iter_rows(min_row=2, max_row=average_row, max_col=5):
        for cell in row:
            cell.font = Font(size=12)
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    for row_idx, session in enumerate(sessions, start=first_data_row):
        if session["duration"] > LONG_LOAD:
            for cell in ws[row_idx][0:5]:
                cell.fill = yellow_fill

    for row_idx in range(first_data_row, average_row + 1):
        ws.cell(row_idx, 2).number_format = "mmm d, yyyy"
        ws.cell(row_idx, 3).number_format = "h:mm:ss AM/PM"
        ws.cell(row_idx, 4).number_format = "h:mm:ss AM/PM"
        ws.cell(row_idx, 5).number_format = "[h]:mm"

    ws.cell(average_row, 4).alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 16
    ws.row_dimensions[1].height = 24
    for row_idx in range(2, average_row + 1):
        ws.row_dimensions[row_idx].height = 21
    ws.freeze_panes = "A2"


def save_single_zone_workbook(zone_config: dict, sessions: list[dict]) -> Path:
    wb = Workbook()
    del wb[wb.active.title]
    write_zone_sheet(wb, zone_config["sheet"], sessions)
    output = OUTPUT_DIR / f"{zone_config['title']} Loading Times for {display_date(REPORT_DAY)}.xlsx"
    wb.save(output)
    return output


def duration_text(value: timedelta) -> str:
    total_minutes = int(value.total_seconds() // 60)
    return f"{total_minutes // 60}:{total_minutes % 60:02d}"


def create_combined_pdf(zone_sessions: list[tuple[dict, list[dict]]]) -> Path:
    styles = getSampleStyleSheet()
    doc = SimpleDocTemplate(
        str(OUTPUT_PDF),
        pagesize=letter,
        leftMargin=0.65 * inch,
        rightMargin=0.65 * inch,
        topMargin=0.45 * inch,
        bottomMargin=0.45 * inch,
    )
    story = []
    for index, (zone_config, sessions) in enumerate(zone_sessions):
        title = Paragraph(
            f"{zone_config['label']} Loading Times - {REPORT_DAY.strftime('%B')} {REPORT_DAY.day}, {REPORT_DAY.year}",
            styles["Title"],
        )
        story.extend([title, Spacer(1, 0.2 * inch)])
        rows = [["Device", "Date", "Start Time", "Stop Time", "Duration"]]
        for session in sessions:
            rows.append(
                [
                    session["device"],
                    session["start"].strftime("%b %d, %Y").replace(" 0", " "),
                    session["start"].strftime("%-I:%M:%S %p"),
                    session["end"].strftime("%-I:%M:%S %p"),
                    duration_text(session["duration"]),
                ]
            )
        average = timedelta(seconds=sum(item["duration"].total_seconds() for item in sessions) / len(sessions)) if sessions else timedelta(0)
        rows.append(["", "", "", "Average", duration_text(average) if sessions else "-"])
        table = PdfTable(
            rows,
            colWidths=[1.75 * inch, 1.25 * inch, 1.45 * inch, 1.45 * inch, 0.85 * inch],
            repeatRows=1,
        )
        commands = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ALIGN", (1, 1), (-1, -1), "CENTER"),
            ("ALIGN", (0, 0), (-1, 0), "LEFT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D9D9D9")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#FAFAFA")]),
            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#E7F0F7")),
            ("TOPPADDING", (0, 0), (-1, -1), 7),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
        ]
        for row_index, session in enumerate(sessions, start=1):
            if session["duration"] > LONG_LOAD:
                commands.append(("BACKGROUND", (0, row_index), (-1, row_index), colors.yellow))
        # Reapply the header styling last so ReportLab preserves it when a
        # short second-plant table follows an explicit page break.
        commands.extend(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ]
        )
        table.setStyle(TableStyle(commands))
        story.append(table)
        if index < len(zone_sessions) - 1:
            story.append(PageBreak())
    doc.build(story)
    return OUTPUT_PDF


def email_output(file_path: Path) -> None:
    host = os.environ.get("SMTP_HOST")
    port = int(os.environ.get("SMTP_PORT") or "587")
    username = os.environ.get("SMTP_USERNAME")
    password = os.environ.get("SMTP_PASSWORD")
    sender = os.environ.get("EMAIL_FROM") or username
    recipients = [item.strip() for item in os.environ.get("EMAIL_TO", "").split(",") if item.strip()]

    if not all([host, username, password, sender]) or not recipients:
        message = "Email skipped: SMTP_HOST, SMTP_USERNAME, SMTP_PASSWORD, EMAIL_FROM/SMTP_USERNAME, and EMAIL_TO are required."
        if os.environ.get("CI") or os.environ.get("REQUIRE_EMAIL") == "1":
            raise SystemExit(message)
        print(message)
        return

    subject_date = f"{display_date(REPORT_DAY)}, {REPORT_DAY.year}"
    message = EmailMessage()
    message["From"] = sender
    message["To"] = ", ".join(recipients)
    message["Subject"] = f"Dixon/Tracy Loading Times for {subject_date}"
    message.set_content(
        "Attached is the combined Dixon and Tracy loading time report for "
        f"{subject_date}.\n\nThis was generated automatically from Geotab."
    )
    message.add_attachment(file_path.read_bytes(), maintype="application", subtype="pdf", filename=file_path.name)

    with smtplib.SMTP(host, port, timeout=60) as smtp:
        smtp.starttls()
        smtp.login(username, password)
        smtp.send_message(message)
    print(f"emailed={file_path}")


def validate_outputs(audit: dict, plant_files: list[Path]) -> None:
    missing_files = [str(path) for path in plant_files if not path.exists() or path.stat().st_size == 0]
    if missing_files:
        raise SystemExit(f"Missing or empty report files: {', '.join(missing_files)}")

    for zone_label, zone_audit in audit["zones"].items():
        if zone_audit["cleanedLoadCount"] < 0:
            raise SystemExit(f"{zone_label} has an invalid cleaned load count.")
        if zone_audit["ignoredUnder25Count"] < 0:
            raise SystemExit(f"{zone_label} has an invalid ignored stop count.")
        if zone_audit["longLoadCount"] > zone_audit["cleanedLoadCount"]:
            raise SystemExit(f"{zone_label} long-load count exceeds cleaned load count.")
        if not zone_audit["startsWithinOperatingWindow"]:
            raise SystemExit(f"{zone_label} contains a session starting outside 5:00 AM-5:00 PM.")
        if zone_audit["maxDurationSeconds"] > MAX_DURATION.total_seconds():
            raise SystemExit(f"{zone_label} contains a duration over 4 hours.")


def main() -> int:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_AUDIT_DIR.mkdir(parents=True, exist_ok=True)
    client = GeotabClient(os.environ.get("GEOTAB_SERVER", "my.geotab.com"), timeout=int(os.environ.get("GEOTAB_TIMEOUT", "120")))
    client.authenticate(require_env("GEOTAB_USERNAME"), require_env("GEOTAB_PASSWORD"), require_env("GEOTAB_DATABASE"))
    events_by_zone, _devices, rule_names_by_zone = fetch_inputs(client)

    wb = Workbook()
    del wb[wb.active.title]
    audit = {
        "reportDate": REPORT_DAY.isoformat(),
        "weekday": REPORT_DAY.strftime("%A"),
        "method": "Pair Geotab zone ENTRY ExceptionEvent activeFrom with the following EXIT ExceptionEvent activeFrom for each device; include entries from 5:00 AM through 4:59 PM without capping exit time; suppress immediate GPS exit/re-entry blips under 25 minutes; exclude visits under 25 minutes and probable errors over 4 hours",
        "rawEventCount": sum(len(events) for events in events_by_zone.values()),
        "rules": rule_names_by_zone,
        "files": {},
        "zones": {},
    }
    plant_files = []
    zone_sessions = []

    for zone_config in ZONES:
        zone_events = events_by_zone[zone_config["label"]]
        sessions, ignored = sessions_for_zone(zone_events)
        zone_sessions.append((zone_config, sessions))
        write_zone_sheet(wb, zone_config["sheet"], sessions)
        single_output = save_single_zone_workbook(zone_config, sessions)
        plant_files.append(single_output)
        sorted_start_times = [item["start"] for item in sessions] == sorted(item["start"] for item in sessions)
        audit["files"][zone_config["label"]] = str(single_output)
        audit["zones"][zone_config["label"]] = {
            "rawEventCount": len(zone_events),
            "cleanedLoadCount": len(sessions),
            "ignoredUnder25Count": sum(
                1
                for item in ignored
                if "under 25 minutes" in item["reason"] or "Immediate re-entry" in item["reason"]
            ),
            "longLoadCount": sum(1 for item in sessions if item["duration"] > LONG_LOAD),
            "probableErrorCount": sum(1 for item in ignored if item["reason"].startswith("Probable error")),
            "averageSeconds": sum(item["duration"].total_seconds() for item in sessions) / len(sessions) if sessions else 0,
            "maxDurationSeconds": max((item["duration"].total_seconds() for item in sessions), default=0),
            "startsWithinOperatingWindow": all(
                OPERATING_START <= item["start"].time() < OPERATING_END
                for item in sessions
            ),
            "sortedStartTimes": sorted_start_times,
        }
        if not sorted_start_times:
            raise SystemExit(f"{zone_config['label']} start times are not sorted.")

    wb.save(OUTPUT_XLSX)
    pdf_output = create_combined_pdf(zone_sessions)
    audit["files"]["combinedPdf"] = str(pdf_output)
    validate_outputs(audit, [OUTPUT_XLSX, *plant_files, pdf_output])
    OUTPUT_JSON.write_text(json.dumps(audit, indent=2), encoding="utf-8")
    print(f"output={OUTPUT_XLSX}")
    print(json.dumps(audit["zones"], indent=2))
    email_output(pdf_output)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
